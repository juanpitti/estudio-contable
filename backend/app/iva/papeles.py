"""Generador de papeles de trabajo Excel para pre-liquidación IVA.

Produce un workbook con 3 hojas:
- Resumen: débito/crédito por alícuota, saldo a pagar o a favor.
- Ventas: detalle de comprobantes de venta.
- Compras: detalle de comprobantes de compra.
"""

from io import BytesIO
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font

from app.iva.calculadora import LiquidacionIva
from app.iva.comprobante import ComprobanteIva


def generar_papel_trabajo(
    liq: LiquidacionIva,
    comprobantes: list[ComprobanteIva],
    periodo: str,
) -> bytes:
    wb = openpyxl.Workbook()

    # Hoja Resumen
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Pre-liquidación IVA"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Período:"
    ws["B3"] = periodo
    ws["A5"] = "DÉBITO FISCAL (ventas)"
    ws["A5"].font = Font(bold=True)
    fila = 6
    for alic, total in sorted(liq.debito.items()):
        ws[f"A{fila}"] = f"Alícuota {alic}"
        ws[f"B{fila}"] = float(total)
        ws[f"B{fila}"].number_format = '"$"#,##0.00'
        fila += 1
    ws[f"A{fila}"] = "Total débito"
    ws[f"A{fila}"].font = Font(bold=True)
    ws[f"B{fila}"] = float(liq.total_debito)
    ws[f"B{fila}"].font = Font(bold=True)
    ws[f"B{fila}"].number_format = '"$"#,##0.00'

    fila += 2
    ws[f"A{fila}"] = "CRÉDITO FISCAL (compras)"
    ws[f"A{fila}"].font = Font(bold=True)
    fila += 1
    for alic, total in sorted(liq.credito.items()):
        ws[f"A{fila}"] = f"Alícuota {alic}"
        ws[f"B{fila}"] = float(total)
        ws[f"B{fila}"].number_format = '"$"#,##0.00'
        fila += 1
    ws[f"A{fila}"] = "Total crédito"
    ws[f"A{fila}"].font = Font(bold=True)
    ws[f"B{fila}"] = float(liq.total_credito)
    ws[f"B{fila}"].font = Font(bold=True)
    ws[f"B{fila}"].number_format = '"$"#,##0.00'

    fila += 2
    if liq.saldo_favor_anterior > 0:
        ws[f"A{fila}"] = "Saldo a favor anterior"
        ws[f"B{fila}"] = float(liq.saldo_favor_anterior)
        ws[f"B{fila}"].number_format = '"$"#,##0.00'
        fila += 1

    if liq.saldo_a_pagar > 0:
        ws[f"A{fila}"] = "SALDO A PAGAR"
        ws[f"A{fila}"].font = Font(bold=True, color="FF0000")
        ws[f"B{fila}"] = float(liq.saldo_a_pagar)
        ws[f"B{fila}"].font = Font(bold=True, color="FF0000")
        ws[f"B{fila}"].number_format = '"$"#,##0.00'
    else:
        ws[f"A{fila}"] = "SALDO A FAVOR (IVA técnico)"
        ws[f"A{fila}"].font = Font(bold=True, color="008000")
        ws[f"B{fila}"] = float(liq.saldo_a_favor_final)
        ws[f"B{fila}"].font = Font(bold=True, color="008000")
        ws[f"B{fila}"].number_format = '"$"#,##0.00'

    # Ajustar anchos
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

    # Hoja Ventas
    ws_v = wb.create_sheet("Ventas")
    ws_v.append(["ID", "Fecha", "Alícuota", "Neto", "IVA", "Confirmado por"])
    for c in sorted((c for c in comprobantes if c.tipo == "venta"), key=lambda x: x.fecha):
        for l in c.lineas:
            ws_v.append([
                c.id, c.fecha.isoformat(), str(l.alicuota),
                float(l.neto), float(l.iva), c.confirmado_por,
            ])
    for col in ["D", "E"]:
        for cell in ws_v[col][1:]:
            cell.number_format = '"$"#,##0.00'

    # Hoja Compras
    ws_c = wb.create_sheet("Compras")
    ws_c.append(["ID", "Fecha", "Alícuota", "Neto", "IVA", "Confirmado por"])
    for c in sorted((c for c in comprobantes if c.tipo == "compra"), key=lambda x: x.fecha):
        for l in c.lineas:
            ws_c.append([
                c.id, c.fecha.isoformat(), str(l.alicuota),
                float(l.neto), float(l.iva), c.confirmado_por,
            ])
    for col in ["D", "E"]:
        for cell in ws_c[col][1:]:
            cell.number_format = '"$"#,##0.00'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
