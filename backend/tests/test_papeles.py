from decimal import Decimal
from datetime import date
from io import BytesIO

import openpyxl

from app.iva.papeles import generar_papel_trabajo
from app.iva.calculadora import liquidacion_iva
from app.iva.comprobante import ComprobanteIva, AlicuotaLinea


def _comp(tipo, fecha, alicuota, neto, iva, id=1):
    return ComprobanteIva(
        id=id, cliente_id=1, tipo=tipo, fecha=date.fromisoformat(fecha),
        lineas=[AlicuotaLinea(Decimal(alicuota), Decimal(neto), Decimal(iva))],
        confirmado_por="test", confirmado_en=None,
    )


def test_genera_excel_con_estructura_correcta():
    ventas = [_comp("venta", "2026-08-01", "0.21", "100000", "21000", id=1)]
    compras = [_comp("compra", "2026-08-01", "0.21", "50000", "10500", id=2)]
    liq = liquidacion_iva(ventas, compras, Decimal("0"))
    data = generar_papel_trabajo(liq, ventas + compras, periodo="2026-08")
    wb = openpyxl.load_workbook(BytesIO(data))
    assert "Resumen" in wb.sheetnames
    assert "Ventas" in wb.sheetnames
    assert "Compras" in wb.sheetnames
    ws = wb["Resumen"]
    assert ws["A1"].value == "Pre-liquidación IVA"
    assert ws["A3"].value == "Período:"
    assert ws["B3"].value == "2026-08"


def test_excel_con_saldo_a_pagar():
    ventas = [_comp("venta", "2026-08-01", "0.21", "100000", "21000", id=1)]
    compras = [_comp("compra", "2026-08-01", "0.21", "50000", "10500", id=2)]
    liq = liquidacion_iva(ventas, compras, Decimal("0"))
    data = generar_papel_trabajo(liq, ventas + compras, periodo="2026-08")
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb["Resumen"]
    # Buscar celda con "SALDO A PAGAR"
    found = False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value and "A PAGAR" in str(cell.value):
                found = True
    assert found


def test_excel_con_saldo_a_favor():
    ventas = [_comp("venta", "2026-08-01", "0.21", "100000", "21000", id=1)]
    compras = [_comp("compra", "2026-08-01", "0.21", "200000", "42000", id=2)]
    liq = liquidacion_iva(ventas, compras, Decimal("0"))
    data = generar_papel_trabajo(liq, ventas + compras, periodo="2026-08")
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb["Resumen"]
    found = False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value and "A FAVOR" in str(cell.value):
                found = True
    assert found
